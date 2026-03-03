#!/usr/bin/env python3
"""
Generate Luther Bibel 1912 SD card directory tree for Flipper Zero.

Output structure:
  luther1912/
    <Section>/
      <Book>/
        <Chapter>/
          verse1.txt
          verse2.txt
          ...

Run:
  python3 generate_sd.py BBLTitles.xlsx BBLgerman.xlsx

Output:  ./luther1912/   (copy to /ext/apps_data/ on your Flipper SD card)
         ./luther1912_sd.zip  (same, zipped for convenience)
"""

import os
import re
import sys
import zipfile
import openpyxl

# ── Config ─────────────────────────────────────────────────────────────────────
TITLES_FILE  = "BBLTitles.xlsx"
GERMAN_FILE  = "BBLgerman.xlsx"
OUT_DIR      = "luther1912"
ZIP_NAME     = "luther1912_sd.zip"

# Map BibleSection values → folder names (underscores, no spaces, safe for FAT32)
SECTION_MAP = {
    "Old Testament":  "Altes_Testament",
    "Prophets":       "Propheten",
    "New Testament":  "Neues_Testament",
    "Apocrypha":      "Apokryphen",
}

def safe_folder(name: str) -> str:
    """Convert a book name to a FAT32-safe folder name."""
    # Replace spaces with underscores
    s = name.replace(" ", "_")
    # Remove or replace chars illegal on FAT32: \ / : * ? " < > |
    s = re.sub(r'[\\/:*?"<>|]', "", s)
    # Collapse multiple underscores
    s = re.sub(r'_+', "_", s).strip("_")
    return s

def load_titles(path: str) -> dict:
    """Returns {IdBook: (section_folder, book_folder, display_name)}"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    books = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_book, _, section, _, _, _, ger_name = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if id_book is None:
            continue
        section_folder = SECTION_MAP.get(section, safe_folder(section))
        book_folder    = safe_folder(str(ger_name))
        books[int(id_book)] = (section_folder, book_folder, str(ger_name))
    return books

def load_verses(path: str):
    """Yields (book_id, chapter, verse, text) tuples — uses the 'scripture' column (clean text)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue  # skip header
        _, book_id, chapter, verse, scripture = row[0], row[1], row[2], row[3], row[4]
        if book_id is None or scripture is None:
            continue
        yield int(book_id), int(chapter), int(verse), str(scripture).strip()

def main():
    titles_path = TITLES_FILE
    german_path = GERMAN_FILE

    print(f"Loading book titles from {titles_path} ...")
    books = load_titles(titles_path)
    print(f"  {len(books)} books found.")

    print(f"Loading verses from {german_path} ...")
    verse_count = 0
    file_count  = 0
    dirs_made   = set()

    with zipfile.ZipFile(ZIP_NAME, "w", zipfile.ZIP_DEFLATED) as zf:
        for book_id, chapter, verse, text in load_verses(german_path):
            if book_id not in books:
                print(f"  WARNING: BookID {book_id} not found in titles, skipping.")
                continue

            section_folder, book_folder, _ = books[book_id]

            # Build directory path
            chapter_dir = os.path.join(OUT_DIR, section_folder, book_folder, str(chapter))
            verse_file  = os.path.join(chapter_dir, f"verse{verse}.txt")

            # Create on disk
            if chapter_dir not in dirs_made:
                os.makedirs(chapter_dir, exist_ok=True)
                dirs_made.add(chapter_dir)

            with open(verse_file, "w", encoding="utf-8") as f:
                f.write(text)

            # Also add to zip (path inside zip mirrors SD card root)
            zip_path = verse_file.replace(OUT_DIR + os.sep, OUT_DIR + "/")
            zf.write(verse_file, zip_path)

            verse_count += 1
            file_count  += 1

            if verse_count % 5000 == 0:
                print(f"  {verse_count} verses written...")

    print(f"\nDone!")
    print(f"  Verses written : {verse_count:,}")
    print(f"  Files created  : {file_count:,}")
    print(f"  Directories    : {len(dirs_made):,}")
    print(f"  Output folder  : ./{OUT_DIR}/")
    print(f"  ZIP archive    : ./{ZIP_NAME}")
    print()
    print("Copy the 'luther1912' folder to your Flipper SD card:")
    print("  /ext/apps_data/luther1912/")

if __name__ == "__main__":
    main()
